"""Tkinter user interface for Archive Search V11."""

from __future__ import annotations

import queue
import shlex
import threading
import tkinter as tk
import tkinter.font as tkfont
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl

from archive_config import (
    APP_TITLE,
    APP_ICON_PATH,
    COMPANY_NAME,
    COMPANY_LOGO_PATH,
    HEADER_LOGO_MAX_HEIGHT,
    HEADER_LOGO_MAX_WIDTH,
    ARCHIVE_SEARCH_PRIMARY_FOLDER,
    ARCHIVE_SEARCH_SECONDARY_FOLDER,
    TEXT_DOCUMENT_EXTENSIONS,
    WORKBOOK_EXTENSIONS,
    BRAND_BLUE,
    BRAND_BLACK,
    BRAND_WHITE,
    BRAND_ORANGE,
    BRAND_GREY,
    NCRT_BLUE,
    APP_BACKGROUND,
    SURFACE_BLUE,
    MUTED_TEXT,
    SOFT_BLUE,
    ROW_BLUE,
    HOVER_BLUE,
    SELECTED_BLUE,
    HIGHLIGHT_TAG_NAME,
    HIGHLIGHT_BACKGROUND,
    HIGHLIGHT_FOREGROUND
)

from archive_extractor import (
    DocumentExtractor,
    HAS_PYXLSB,
    HAS_XLRD,
    convert_xls_value,
    open_xlsb_workbook,
    xlrd,
)
from archive_index import SQLiteIndexManager
from archive_models import SearchResult, SearchRoot
from archive_opener import FileOpener
from archive_runner import SearchRunner
from archive_utils import (
    build_highlight_terms,
    display_text,
    get_app_folder,
    infer_type_label,
    pluralize,
)


class ArchiveSearchApp:
    """Main Tkinter application."""

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.tray_icon = None
        self.tray_icon_thread: Optional[threading.Thread] = None
        self.app_icon_photo = None
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self._set_initial_window_geometry()
        self._configure_window_icons()
        self._start_system_tray_icon()

        self.app_folder = get_app_folder()
        # Worker threads never touch Tkinter directly; they post messages here
        # and the UI thread consumes them in _poll_queue().
        self.queue: queue.Queue = queue.Queue()
        self.cancel_event = threading.Event()
        self.search_thread: Optional[threading.Thread] = None
        self.index_status_thread: Optional[threading.Thread] = None
        self.index_status_cancel_event = threading.Event()
        self.current_job_mode = "search"

        self.extractor = DocumentExtractor()
        self.index_manager = SQLiteIndexManager(self.extractor)

        self.current_highlight_terms: List[str] = []
        self.results_by_iid: Dict[str, SearchResult] = {}
        self.result_heading_labels = {
            "document": "File",
            "type": "Type",
            "matches": "Matches",
        }
        self.result_sort_column: Optional[str] = None
        self.result_sort_reverse = False
        self.hovered_result_iid: Optional[str] = None

        self.primary_folder_var = tk.StringVar(value=str(ARCHIVE_SEARCH_PRIMARY_FOLDER))
        self.secondary_folder_var = tk.StringVar(value=str(ARCHIVE_SEARCH_SECONDARY_FOLDER))
        self.use_primary_var = tk.BooleanVar(value=True)
        self.use_secondary_var = tk.BooleanVar(value=True)
        self.primary_subfolders_var = tk.BooleanVar(value=True)
        self.secondary_subfolders_var = tk.BooleanVar(value=True)

        self.search_workbooks_var = tk.BooleanVar(value=True)
        self.search_text_documents_var = tk.BooleanVar(value=True)

        self.search_var = tk.StringVar()
        self.search_placeholder_text = "Search by table model, machine, customer, order number…"
        self.search_placeholder_active = False
        self.search_placeholder_changing = False
        self.match_mode_var = tk.StringVar(value="any")
        self.status_var = tk.StringVar(value="Ready.")
        self.result_summary_var = tk.StringVar(value="No results yet")
        self.index_info_var = tk.StringVar(value=self.index_manager.get_index_path())
        self.index_status_var = tk.StringVar(value="Index status: not checked yet.")

        # Folder and file-type changes affect the displayed index paths.
        for variable in (self.primary_folder_var, self.secondary_folder_var):
            variable.trace_add("write", lambda *_args: self._on_index_settings_changed())
        for variable in (
            self.use_primary_var,
            self.use_secondary_var,
            self.primary_subfolders_var,
            self.secondary_subfolders_var,
            self.search_workbooks_var,
            self.search_text_documents_var,
        ):
            variable.trace_add("write", lambda *_args: self._on_index_settings_changed())

        self._build_ui()
        self.root.bind("<Configure>", self._on_window_configure)
        self._update_status_wraplengths()
        self._update_folder_states()
        self._refresh_index_info()
        self._poll_queue()

    def _set_initial_window_geometry(self) -> None:
        """Open the app centred on the current screen and avoid oversized windows."""
        target_width = 1480
        target_height = 900
        try:
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            width = min(target_width, max(900, screen_width - 80))
            height = min(target_height, max(650, screen_height - 120))
            x = max(0, (screen_width - width) // 2)
            y = max(0, (screen_height - height) // 2)
            self.root.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            self.root.geometry(f"{target_width}x{target_height}")

    def _build_ui(self) -> None:
        self._configure_styles()
        self.advanced_visible = False

        container = ttk.Frame(self.root, padding=0, style="App.TFrame")
        container.pack(fill="both", expand=True)

        header = tk.Frame(container, bg=BRAND_WHITE, highlightthickness=0)
        header.pack(fill="x")
        header.columnconfigure(1, weight=1)

        accent_bar = tk.Frame(header, bg=BRAND_BLUE, width=7)
        accent_bar.grid(row=0, column=0, sticky="ns")

        header_text = tk.Frame(header, bg=BRAND_WHITE)
        header_text.grid(row=0, column=1, sticky="we", padx=(22, 18), pady=(14, 13))
        self._create_header_brand_mark(header_text).pack(anchor="w")
        tk.Label(
            header_text,
            text="Archive Search",
            bg=BRAND_WHITE,
            fg=BRAND_BLUE,
            font=self._ui_font("heading", 21),
            anchor="w",
        ).pack(anchor="w", pady=(3, 0))
        tk.Label(
            header_text,
            text="Find previous rotary-table supply experience by table model, machine, customer, order number, or other information.",
            bg=BRAND_WHITE,
            fg=MUTED_TEXT,
            font=self._ui_font("body", 9),
            anchor="w",
        ).pack(anchor="w", pady=(3, 0))

        self.advanced_toggle_button = ttk.Button(
            header,
            text="Archive settings",
            command=self._toggle_archive_settings,
            style="Header.TButton",
        )
        self.advanced_toggle_button.grid(row=0, column=2, sticky="e", padx=(0, 18), pady=(14, 13))

        content = ttk.Frame(container, padding=(16, 12, 16, 12), style="App.TFrame")
        content.pack(fill="both", expand=True)

        self.search_card = ttk.Frame(content, padding=(16, 12), style="Card.TFrame")
        search_card = self.search_card
        search_card.pack(fill="x")
        search_card.columnconfigure(0, weight=1)

        top_search_row = ttk.Frame(search_card, style="Surface.TFrame")
        top_search_row.grid(row=0, column=0, columnspan=4, sticky="we")
        top_search_row.columnconfigure(1, weight=1)

        ttk.Label(
            top_search_row,
            text="Search the archive",
            font=self._ui_font("heading", 12),
            style="Card.TLabel",
        ).grid(row=0, column=0, sticky="w")

        mode_row = ttk.Frame(top_search_row, style="Surface.TFrame")
        mode_row.grid(row=0, column=1, sticky="e")
        ttk.Label(mode_row, text="Search mode:", style="Muted.Card.TLabel").pack(side="left")
        ttk.Radiobutton(mode_row, text="Any word", variable=self.match_mode_var, value="any").pack(side="left", padx=(8, 0))
        ttk.Radiobutton(mode_row, text="All words", variable=self.match_mode_var, value="all").pack(side="left", padx=(8, 0))

        self.search_entry = ttk.Entry(search_card, textvariable=self.search_var, width=84, style="Search.TEntry")
        self.search_entry.grid(row=1, column=0, sticky="we", padx=(0, 10), pady=(8, 0), ipady=4)
        self.search_placeholder_label = tk.Label(
            self.search_entry,
            text=self.search_placeholder_text,
            bg=BRAND_WHITE,
            fg=MUTED_TEXT,
            font=self._ui_font("body", 9),
            anchor="w",
            cursor="xterm",
        )
        self.search_placeholder_label.bind("<Button-1>", self._focus_empty_search_entry)
        self.search_placeholder_label.bind("<ButtonRelease-1>", self._focus_empty_search_entry)
        self.search_entry.bind("<Return>", lambda event: self.start_search())
        self.search_entry.bind("<FocusIn>", self._on_search_entry_focus_changed)
        self.search_entry.bind("<FocusOut>", self._on_search_entry_focus_changed)
        self.search_var.trace_add("write", lambda *_args: self._on_search_text_changed())
        self._show_search_placeholder()

        self.search_button = ttk.Button(search_card, text="Search archive", command=self.start_search, style="Accent.TButton")
        self.search_button.grid(row=1, column=1, sticky="e", pady=(8, 0))
        self.clear_button = ttk.Button(search_card, text="Clear results", command=self.clear_output)
        self.clear_button.grid(row=1, column=2, sticky="e", padx=(8, 0), pady=(8, 0))
        self.cancel_button = ttk.Button(search_card, text="Cancel", command=self.cancel_process, state="disabled")
        self.cancel_button.grid(row=1, column=3, sticky="e", padx=(8, 0), pady=(8, 0))

        index_panel = ttk.Frame(search_card, padding=(10, 7), style="Info.TFrame")
        index_panel.grid(row=2, column=0, columnspan=4, sticky="we", pady=(9, 0))
        index_panel.columnconfigure(1, weight=1)
        ttk.Label(index_panel, text="Index status", font=self._ui_font("body", 9, "bold"), style="Info.TLabel").grid(
            row=0, column=0, sticky="w", padx=(0, 10)
        )
        self.index_status_label = tk.Label(
            index_panel,
            textvariable=self.index_status_var,
            anchor="w",
            justify="left",
            fg=BRAND_BLACK,
            bg=SURFACE_BLUE,
            font=self._ui_font("body", 9),
        )
        self.index_status_label.grid(row=0, column=1, sticky="we")
        self.check_index_button = ttk.Button(
            index_panel,
            text="Check index",
            command=self.check_index_status,
            style="Secondary.TButton",
        )
        self.check_index_button.grid(row=0, column=2, sticky="e", padx=(12, 0))

        self.status_label = tk.Label(
            search_card,
            textvariable=self.status_var,
            anchor="nw",
            justify="left",
            bg=BRAND_WHITE,
            fg=MUTED_TEXT,
            font=self._ui_font("body", 9),
        )
        self.status_label.grid(row=3, column=0, columnspan=4, sticky="we", pady=(8, 0))

        self.progress_bar = ttk.Progressbar(search_card, mode="indeterminate", style="Brand.Horizontal.TProgressbar")
        self.progress_bar.grid(row=4, column=0, columnspan=4, sticky="we", pady=(8, 0))
        self.progress_bar.grid_remove()

        self.archive_settings_frame = ttk.Frame(content, padding=(14, 12), style="Card.TFrame")
        self._build_archive_settings_panel(self.archive_settings_frame)

        results_header = ttk.Frame(content, style="App.TFrame")
        results_header.pack(fill="x", pady=(12, 6))
        ttk.Label(results_header, text="Search results", font=self._ui_font("heading", 12), style="App.TLabel").pack(side="left")
        ttk.Label(results_header, textvariable=self.result_summary_var, style="Muted.App.TLabel").pack(side="left", padx=(12, 0))
        ttk.Label(
            results_header,
            text="Select a row to review matches. Double-click a result to open it; double-click a column heading or divider to auto-fit its width.",
            style="Muted.App.TLabel",
        ).pack(side="left", padx=(12, 0))
        self.open_button = ttk.Button(
            results_header,
            text="Open selected",
            command=self.open_selected_result,
            style="Secondary.TButton",
            state="disabled",
        )
        self.open_button.pack(side="right")
        self.show_folder_button = ttk.Button(
            results_header,
            text="Show in folder",
            command=self.reveal_selected_result,
            state="disabled",
        )
        self.show_folder_button.pack(side="right", padx=(0, 8))
        self.copy_path_button = ttk.Button(results_header, text="Copy path", command=self.copy_selected_path, state="disabled")
        self.copy_path_button.pack(side="right", padx=(0, 8))

        self.panes = ttk.Panedwindow(content, orient="horizontal")
        panes = self.panes
        panes.pack(fill="both", expand=True)
        self.root.after_idle(self._set_initial_pane_sash)

        left_frame = ttk.Frame(panes, style="Card.TFrame", padding=8)
        right_frame = ttk.Frame(panes, style="Card.TFrame", padding=8)
        panes.add(left_frame, weight=2)
        panes.add(right_frame, weight=8)

        columns = ("document", "type", "matches")
        self.results_tree = ttk.Treeview(left_frame, columns=columns, show="headings", selectmode="browse", style="Results.Treeview")
        self._update_result_headings()

        self.results_tree.column("document", width=300, anchor="w", minwidth=150)
        self.results_tree.column("type", width=72, anchor="center", minwidth=60)
        self.results_tree.column("matches", width=76, anchor="center", minwidth=66)
        self.results_tree.tag_configure("odd", background=BRAND_WHITE)
        self.results_tree.tag_configure("even", background=ROW_BLUE)
        self.results_tree.tag_configure("hover", background=HOVER_BLUE)

        tree_scroll_y = ttk.Scrollbar(left_frame, orient="vertical", command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=tree_scroll_y.set)

        self.results_tree.pack(fill="both", expand=True, side="left")
        tree_scroll_y.pack(fill="y", side="right")

        self.results_tree.bind("<<TreeviewSelect>>", self._on_result_selected)
        self.results_tree.bind("<Double-1>", self._on_results_double_click)
        self.results_tree.bind("<Motion>", self._on_results_mouse_motion)
        self.results_tree.bind("<Leave>", self._on_results_mouse_leave)
        self.results_tree.bind("<Button-3>", self._show_results_context_menu)
        self.results_tree.bind("<Control-Button-1>", self._show_results_context_menu)
        self._build_results_context_menu()

        detail_header = ttk.Frame(right_frame, style="Surface.TFrame")
        detail_header.pack(fill="x", pady=(0, 6))

        self.detail_title_var = tk.StringVar(value="Result details")
        self.detail_meta_var = tk.StringVar(value="Select a result to review the matching evidence found in the archive.")
        ttk.Label(detail_header, textvariable=self.detail_title_var, font=self._ui_font("heading", 12), style="Card.TLabel").pack(anchor="w")
        ttk.Label(detail_header, textvariable=self.detail_meta_var, style="Muted.Card.TLabel").pack(anchor="w", pady=(2, 0))

        detail_text_frame = ttk.Frame(right_frame, style="Surface.TFrame")
        detail_text_frame.pack(fill="both", expand=True)
        detail_text_frame.rowconfigure(0, weight=1)
        detail_text_frame.columnconfigure(0, weight=1)

        self.details_text = tk.Text(
            detail_text_frame,
            wrap="word",
            font=self._ui_font("body", 10),
            borderwidth=0,
            padx=12,
            pady=12,
            bg=BRAND_WHITE,
            fg=BRAND_BLACK,
            insertbackground=BRAND_BLUE,
        )
        detail_scroll_y = ttk.Scrollbar(detail_text_frame, orient="vertical", command=self.details_text.yview)
        detail_scroll_x = ttk.Scrollbar(detail_text_frame, orient="horizontal", command=self.details_text.xview)
        self.details_text.configure(yscrollcommand=detail_scroll_y.set, xscrollcommand=detail_scroll_x.set)

        self.details_text.grid(row=0, column=0, sticky="nsew")
        detail_scroll_y.grid(row=0, column=1, sticky="ns")
        detail_scroll_x.grid(row=1, column=0, sticky="we")

        self.details_text.tag_configure(HIGHLIGHT_TAG_NAME, background=HIGHLIGHT_BACKGROUND, foreground=HIGHLIGHT_FOREGROUND)
        self.details_text.tag_configure("heading", font=self._ui_font("body", 10, "bold"), foreground=BRAND_BLUE)
        self.details_text.tag_configure("muted", foreground=MUTED_TEXT)
        self.details_text.tag_configure("empty_title", font=self._ui_font("heading", 13), foreground=BRAND_BLUE, spacing3=6)
        self.details_text.tag_configure("empty_body", font=self._ui_font("body", 10), foreground=MUTED_TEXT, spacing3=5)
        self.details_text.tag_configure("empty_bullet", font=self._ui_font("body", 10), foreground=BRAND_BLACK, lmargin1=16, lmargin2=16, spacing3=4)
        self.details_text.tag_configure("snippet_label", font=self._ui_font("body", 9, "bold"), foreground=BRAND_BLUE)
        self.details_text.tag_configure("match_title", font=self._ui_font("body", 10, "bold"), foreground=BRAND_BLACK, spacing1=3)
        self.details_text.tag_configure("location", font=self._ui_font("body", 9, "bold"), foreground=MUTED_TEXT)
        self.details_text.tag_configure("snippet_body", font=self._ui_font("body", 10), foreground=BRAND_BLACK, lmargin1=14, lmargin2=14, spacing3=6)
        self.details_text.tag_configure("snippet_separator", foreground=BRAND_GREY)
        self.details_text.tag_configure("workbook_sheet", font=self._ui_font("heading", 10), foreground=BRAND_BLUE, spacing1=8, spacing3=3)
        self.details_text.tag_configure("workbook_header", font=self._ui_font("mono", 10, "bold"), foreground=BRAND_BLUE, background=SOFT_BLUE, spacing1=2, spacing3=2)
        self.details_text.tag_configure("workbook_rule", font=self._ui_font("mono", 10), foreground=MUTED_TEXT, background=SOFT_BLUE)
        self.details_text.tag_configure("workbook_row_odd", font=self._ui_font("mono", 10), foreground=BRAND_BLACK, background=BRAND_WHITE)
        self.details_text.tag_configure("workbook_row_even", font=self._ui_font("mono", 10), foreground=BRAND_BLACK, background=ROW_BLUE)
        self.details_text.tag_raise("sel")
        self.details_text.tag_lower(HIGHLIGHT_TAG_NAME, "sel")
        self.details_text.config(state="disabled")
        self._clear_details()
        self._bind_keyboard_shortcuts()
        self.root.after_idle(lambda: (self.root.focus_set(), self._show_search_placeholder()))


    def _load_icon_photo(self, path: Path, max_size: int = 256) -> Optional[tk.PhotoImage]:
        """Load an application icon for Tk's window/taskbar icon support."""
        try:
            from PIL import Image, ImageTk  # type: ignore
        except Exception:
            return None

        try:
            image = Image.open(path)
            if image.mode not in ("RGBA", "LA"):
                image = image.convert("RGBA")
            image.thumbnail((max_size, max_size))
            return ImageTk.PhotoImage(image)
        except Exception:
            return None

    def _configure_window_icons(self) -> None:
        """Apply the configured local icon to the window title bar and taskbar."""
        if not APP_ICON_PATH or not APP_ICON_PATH.is_file():
            return

        # Windows handles .ico files best via iconbitmap, which controls the
        # title-bar icon and usually the taskbar icon too. Keep this optional so
        # PNG icons and non-Windows systems can still use iconphoto below.
        if APP_ICON_PATH.suffix.lower() == ".ico":
            try:
                self.root.iconbitmap(default=str(APP_ICON_PATH))
            except tk.TclError:
                pass

        self.app_icon_photo = self._load_icon_photo(APP_ICON_PATH)
        if self.app_icon_photo is not None:
            try:
                self.root.iconphoto(True, self.app_icon_photo)
            except tk.TclError:
                pass

    def _start_system_tray_icon(self) -> None:
        """Start an optional system-tray icon when pystray and an icon are available."""
        if not APP_ICON_PATH or not APP_ICON_PATH.is_file():
            return
        try:
            from PIL import Image  # type: ignore
            import pystray  # type: ignore
        except Exception:
            return

        try:
            tray_image = Image.open(APP_ICON_PATH)
        except Exception:
            return

        def show_from_tray(_icon=None, _item=None) -> None:
            self.root.after(0, self._show_window)

        def quit_from_tray(_icon=None, _item=None) -> None:
            self.root.after(0, self._on_close)

        try:
            menu = pystray.Menu(
                pystray.MenuItem("Show Archive Search", show_from_tray, default=True),
                pystray.MenuItem("Quit", quit_from_tray),
            )
            self.tray_icon = pystray.Icon("Archive Search", tray_image, APP_TITLE, menu)
        except Exception:
            self.tray_icon = None
            return

        def run_tray() -> None:
            try:
                self.tray_icon.run()
            except Exception:
                pass

        self.tray_icon_thread = threading.Thread(target=run_tray, daemon=True)
        self.tray_icon_thread.start()

    def _show_window(self) -> None:
        """Bring the main window back to the foreground from the tray menu."""
        try:
            self.root.deiconify()
            self.root.lift()
            self.root.focus_force()
        except tk.TclError:
            pass

    def _stop_system_tray_icon(self) -> None:
        """Stop the optional tray icon before the Tk root is destroyed."""
        icon = getattr(self, "tray_icon", None)
        if icon is None:
            return
        self.tray_icon = None
        try:
            icon.stop()
        except Exception:
            pass

    def _on_close(self) -> None:
        """Close the application and clean up optional background UI integrations."""
        cancel_event = getattr(self, "cancel_event", None)
        if cancel_event is not None:
            cancel_event.set()
        index_status_cancel_event = getattr(self, "index_status_cancel_event", None)
        if index_status_cancel_event is not None:
            index_status_cancel_event.set()
        self._stop_system_tray_icon()
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def _load_logo_with_pillow(self, path: Path) -> Optional[tk.PhotoImage]:
        """Load and scale a header logo with Pillow when it is available."""
        try:
            from PIL import Image, ImageTk  # type: ignore
        except Exception:
            return None

        try:
            image = Image.open(path)
            if image.mode not in ("RGBA", "LA"):
                image = image.convert("RGBA")
            image.thumbnail((HEADER_LOGO_MAX_WIDTH, HEADER_LOGO_MAX_HEIGHT))
            return ImageTk.PhotoImage(image)
        except Exception:
            return None

    def _load_logo_with_tk(self, path: Path) -> Optional[tk.PhotoImage]:
        """Load a PNG/GIF logo using Tk's built-in image support."""
        try:
            photo = tk.PhotoImage(file=str(path))
        except tk.TclError:
            return None

        width = max(1, photo.width())
        height = max(1, photo.height())
        scale = max(
            1,
            (width + HEADER_LOGO_MAX_WIDTH - 1) // HEADER_LOGO_MAX_WIDTH,
            (height + HEADER_LOGO_MAX_HEIGHT - 1) // HEADER_LOGO_MAX_HEIGHT,
        )
        if scale > 1:
            try:
                photo = photo.subsample(scale, scale)
            except tk.TclError:
                return None
        return photo

    def _load_header_logo_photo(self) -> Optional[tk.PhotoImage]:
        """Return a configured local logo image, or None to use company-name text."""
        if not COMPANY_LOGO_PATH or not COMPANY_LOGO_PATH.is_file():
            return None

        photo = self._load_logo_with_pillow(COMPANY_LOGO_PATH)
        if photo is None:
            photo = self._load_logo_with_tk(COMPANY_LOGO_PATH)
        return photo

    def _create_header_brand_mark(self, parent: tk.Widget) -> tk.Label:
        """Show the configured logo image when available, otherwise company-name text."""
        self.header_logo_photo = self._load_header_logo_photo()
        if self.header_logo_photo is not None:
            return tk.Label(
                parent,
                image=self.header_logo_photo,
                bg=BRAND_WHITE,
                bd=0,
                highlightthickness=0,
                anchor="w",
            )

        return tk.Label(
            parent,
            text=COMPANY_NAME.upper(),
            bg=BRAND_WHITE,
            fg=BRAND_BLUE,
            font=self._ui_font("heading", 12),
            anchor="w",
        )

    def _choose_font_family(self, preferred: Sequence[str], fallback: str) -> str:
        """Return the first installed font family from preferred, or fallback."""
        try:
            available = {family.casefold(): family for family in tkfont.families(self.root)}
        except tk.TclError:
            return fallback
        for family in preferred:
            installed = available.get(family.casefold())
            if installed:
                return installed
        return fallback

    def _set_brand_fonts(self) -> None:
        """Prefer Kitagawa's primary typeface, then use the approved Arial fallback."""
        self.body_font_family = self._choose_font_family(("HelveticaNeueLT Std", "Helvetica Neue"), "Arial")
        self.heading_font_family = self._choose_font_family(
            ("HelveticaNeueLT Std Blk", "HelveticaNeueLT Std", "Helvetica Neue"),
            "Arial Black",
        )
        self.light_font_family = self._choose_font_family(
            ("HelveticaNeueLT Std Lt Cn", "HelveticaNeueLT Std", "Helvetica Neue"),
            "Arial",
        )
        self.mono_font_family = self._choose_font_family(("Consolas", "Courier New"), "Courier New")

    def _ui_font(self, role: str, size: int, *modifiers: str) -> Tuple[object, ...]:
        family_by_role = {
            "body": getattr(self, "body_font_family", "Arial"),
            "heading": getattr(self, "heading_font_family", "Arial Black"),
            "light": getattr(self, "light_font_family", "Arial"),
            "mono": getattr(self, "mono_font_family", "Courier New"),
        }
        return (family_by_role.get(role, family_by_role["body"]), size, *modifiers)

    def _configure_styles(self) -> None:
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        self._set_brand_fonts()

        self.root.configure(bg=APP_BACKGROUND)
        style.configure(".", font=self._ui_font("body", 9), foreground=BRAND_BLACK)
        style.configure("App.TFrame", background=APP_BACKGROUND)
        style.configure("App.TLabel", background=APP_BACKGROUND, foreground=BRAND_BLACK)
        style.configure("Muted.App.TLabel", background=APP_BACKGROUND, foreground=MUTED_TEXT)
        style.configure("Card.TFrame", background=BRAND_WHITE, relief="solid", borderwidth=1)
        style.configure("Surface.TFrame", background=BRAND_WHITE)
        style.configure("Info.TFrame", background=SURFACE_BLUE, relief="solid", borderwidth=1)
        style.configure("Card.TLabel", background=BRAND_WHITE, foreground=BRAND_BLACK)
        style.configure("Info.TLabel", background=SURFACE_BLUE, foreground=BRAND_BLACK)
        style.configure("Muted.Card.TLabel", background=BRAND_WHITE, foreground=MUTED_TEXT)
        style.configure("Pill.Card.TLabel", background=SOFT_BLUE, foreground=BRAND_BLUE, padding=(8, 3))
        style.configure("Warning.Card.TLabel", background=BRAND_WHITE, foreground="#6b4200", font=self._ui_font("body", 9, "bold"))
        style.configure("FolderOn.TCheckbutton", font=self._ui_font("body", 10, "bold"), background=BRAND_WHITE, foreground=BRAND_BLACK)
        style.configure("FolderOff.TCheckbutton", font=self._ui_font("body", 10, "bold"), foreground="#808080", background=BRAND_WHITE)
        style.configure("Search.TEntry", fieldbackground=BRAND_WHITE, foreground=BRAND_BLACK, insertcolor=BRAND_BLUE, padding=(9, 7))
        style.configure("TEntry", fieldbackground=BRAND_WHITE, foreground=BRAND_BLACK)
        style.configure("TCheckbutton", background=BRAND_WHITE, foreground=BRAND_BLACK)
        style.configure("TRadiobutton", background=BRAND_WHITE, foreground=BRAND_BLACK)
        style.configure("TButton", padding=(10, 6), background=BRAND_GREY, foreground=BRAND_BLACK)
        style.map(
            "TButton",
            background=[("disabled", "#e5e7eb"), ("active", "#d9dcde")],
            foreground=[("disabled", "#8a8f94")],
        )
        style.configure("Accent.TButton", font=self._ui_font("body", 10, "bold"), padding=(17, 9), background=BRAND_BLUE, foreground=BRAND_WHITE)
        style.map(
            "Accent.TButton",
            background=[("disabled", NCRT_BLUE), ("active", "#004f94")],
            foreground=[("disabled", "#ffffff"), ("active", "#ffffff")],
        )
        style.configure("Secondary.TButton", padding=(12, 6), background=BRAND_WHITE, foreground=BRAND_BLUE)
        style.map(
            "Secondary.TButton",
            background=[("disabled", "#e5e7eb"), ("active", "#e8f2fc")],
            foreground=[("disabled", "#8a8f94"), ("active", BRAND_BLUE)],
        )
        style.configure("Header.TButton", padding=(12, 7), background=BRAND_ORANGE, foreground=BRAND_BLACK)
        style.map(
            "Header.TButton",
            background=[("disabled", "#e5e7eb"), ("active", "#e3a30f")],
            foreground=[("disabled", "#8a8f94"), ("active", BRAND_BLACK)],
        )
        style.configure(
            "Brand.Horizontal.TProgressbar",
            troughcolor=SOFT_BLUE,
            background=BRAND_BLUE,
            bordercolor=BRAND_WHITE,
            lightcolor=BRAND_BLUE,
            darkcolor=BRAND_BLUE,
        )
        style.configure("Results.Treeview", rowheight=30, background=BRAND_WHITE, fieldbackground=BRAND_WHITE, foreground=BRAND_BLACK, borderwidth=0)
        style.configure("Results.Treeview.Heading", font=self._ui_font("body", 9, "bold"), background=BRAND_GREY, foreground=BRAND_BLACK, relief="flat")
        style.map("Results.Treeview", background=[("selected", SELECTED_BLUE)], foreground=[("selected", BRAND_BLACK)])

    def _build_archive_settings_panel(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)

        ttk.Label(parent, text="Archive folders & index", font=self._ui_font("heading", 11), style="Card.TLabel").grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 3)
        )
        ttk.Label(
            parent,
            text="Choose which archives to search and refresh the saved index when new files are added.",
            style="Muted.Card.TLabel",
        ).grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(0, 10)
        )

        self.primary_check = ttk.Checkbutton(
            parent,
            text="Primary folder",
            variable=self.use_primary_var,
            command=self._update_folder_states,
            style="FolderOn.TCheckbutton",
        )
        self.primary_check.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(0, 6))

        self.primary_folder_entry = ttk.Entry(parent, textvariable=self.primary_folder_var, width=90)
        self.primary_folder_entry.grid(row=2, column=1, sticky="we", padx=(0, 8), pady=(0, 6))

        self.primary_browse_button = ttk.Button(
            parent,
            text="Browse...",
            command=lambda: self._browse_folder(self.primary_folder_var, "Select primary search folder"),
        )
        self.primary_browse_button.grid(row=2, column=2, sticky="e", pady=(0, 6))

        self.primary_subfolders_check = ttk.Checkbutton(parent, text="Subfolders", variable=self.primary_subfolders_var)
        self.primary_subfolders_check.grid(row=2, column=3, sticky="w", padx=(8, 0), pady=(0, 6))

        self.secondary_check = ttk.Checkbutton(
            parent,
            text="Secondary folder",
            variable=self.use_secondary_var,
            command=self._update_folder_states,
            style="FolderOn.TCheckbutton",
        )
        self.secondary_check.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=(0, 6))

        self.secondary_folder_entry = ttk.Entry(parent, textvariable=self.secondary_folder_var, width=90)
        self.secondary_folder_entry.grid(row=3, column=1, sticky="we", padx=(0, 8), pady=(0, 6))

        self.secondary_browse_button = ttk.Button(
            parent,
            text="Browse...",
            command=lambda: self._browse_folder(self.secondary_folder_var, "Select secondary search folder"),
        )
        self.secondary_browse_button.grid(row=3, column=2, sticky="e", pady=(0, 6))

        self.secondary_subfolders_check = ttk.Checkbutton(parent, text="Subfolders", variable=self.secondary_subfolders_var)
        self.secondary_subfolders_check.grid(row=3, column=3, sticky="w", padx=(8, 0), pady=(0, 6))

        options = ttk.Frame(parent, style="Surface.TFrame")
        options.grid(row=4, column=0, columnspan=4, sticky="we", pady=(6, 6))
        ttk.Label(options, text="File types:", style="Muted.Card.TLabel").pack(side="left")
        self.workbooks_check = ttk.Checkbutton(options, text="Workbooks", variable=self.search_workbooks_var)
        self.workbooks_check.pack(side="left", padx=(8, 0))
        self.text_documents_check = ttk.Checkbutton(options, text="Text documents", variable=self.search_text_documents_var)
        self.text_documents_check.pack(side="left", padx=(8, 0))

        maintenance = ttk.Frame(parent, style="Surface.TFrame")
        maintenance.grid(row=5, column=0, columnspan=4, sticky="we", pady=(4, 6))
        self.update_button = ttk.Button(maintenance, text="Update index", command=self.update_index, style="Secondary.TButton")
        self.update_button.pack(side="left")

        help_text = "Use Update index after adding, changing, or deleting archive files. Searches use the last saved index."
        ttk.Label(parent, text=help_text, style="Muted.Card.TLabel", wraplength=1100).grid(
            row=6, column=0, columnspan=4, sticky="w", pady=(4, 2)
        )
        update_warning = (
            "Warning: Update index needs exclusive access. It will not work if another user or computer "
            "is currently editing or updating the same index."
        )
        ttk.Label(parent, text=update_warning, style="Warning.Card.TLabel", wraplength=1100).grid(
            row=7, column=0, columnspan=4, sticky="w", pady=(0, 8)
        )

        info_panel = ttk.Frame(parent, style="Surface.TFrame")
        info_panel.grid(row=8, column=0, columnspan=4, sticky="we")

        self.index_info_label = tk.Label(
            info_panel,
            textvariable=self.index_info_var,
            anchor="nw",
            justify="left",
            fg=MUTED_TEXT,
            bg=BRAND_WHITE,
            font=self._ui_font("body", 9),
            height=3,
        )
        self.index_info_label.pack(fill="x", anchor="w")


    def _toggle_archive_settings(self) -> None:
        self.advanced_visible = not self.advanced_visible
        if self.advanced_visible:
            self.archive_settings_frame.pack(fill="x", pady=(12, 0), after=self.search_card)
            self.advanced_toggle_button.config(text="Hide settings")
        else:
            self.archive_settings_frame.pack_forget()
            self.advanced_toggle_button.config(text="Archive settings")

    def _bind_keyboard_shortcuts(self) -> None:
        """Add a few common shortcuts without taking over normal text entry."""
        self.root.bind("<Control-f>", self._focus_search)
        self.root.bind("<Control-F>", self._focus_search)
        self.root.bind("<Control-l>", self._focus_search)
        self.root.bind("<Control-L>", self._focus_search)
        self.root.bind("<Control-o>", self._open_selected_from_shortcut)
        self.root.bind("<Control-O>", self._open_selected_from_shortcut)
        self.root.bind("<Escape>", self._cancel_from_shortcut)

    def _focus_search(self, event=None):
        self.search_entry.focus_set()
        self._hide_search_placeholder()
        if self.search_var.get():
            self.search_entry.select_range(0, "end")
            self.search_entry.icursor("end")
        return "break"

    def _open_selected_from_shortcut(self, event=None):
        if self.results_tree.selection():
            self.open_selected_result()
        return "break"

    def _cancel_from_shortcut(self, event=None):
        if self.search_thread and self.search_thread.is_alive():
            self.cancel_process()
            return "break"
        return None

    def _set_initial_pane_sash(self) -> None:
        """Give the details pane slightly more starting width than the results pane."""
        panes = getattr(self, "panes", None)
        if panes is None:
            return
        try:
            width = panes.winfo_width()
            if width > 200:
                panes.sashpos(0, int(width * 0.31))
        except Exception:
            pass

    def _on_index_settings_changed(self) -> None:
        self._refresh_index_info()
        self.index_status_var.set("Index status: settings changed. Click Check index when needed.")

    def _collect_selected_search_roots(self) -> Tuple[SearchRoot, ...]:
        """Return only enabled root folders that currently exist on disk."""
        roots: List[SearchRoot] = []
        if self.use_primary_var.get():
            primary_text = self.primary_folder_var.get().strip()
            if primary_text:
                primary_path = Path(primary_text).expanduser()
                if primary_path.exists() and primary_path.is_dir():
                    roots.append(SearchRoot("Primary", primary_path, self.primary_subfolders_var.get()))
        if self.use_secondary_var.get():
            secondary_text = self.secondary_folder_var.get().strip()
            if secondary_text:
                secondary_path = Path(secondary_text).expanduser()
                if secondary_path.exists() and secondary_path.is_dir():
                    roots.append(SearchRoot("Secondary", secondary_path, self.secondary_subfolders_var.get()))
        return tuple(roots)

    def _collect_allowed_extensions(self) -> Tuple[str, ...]:
        """Translate the UI checkboxes into the suffixes the index should scan."""
        allowed_extensions = set()
        if self.search_workbooks_var.get():
            allowed_extensions |= WORKBOOK_EXTENSIONS
        if self.search_text_documents_var.get():
            allowed_extensions |= TEXT_DOCUMENT_EXTENSIONS
        return tuple(sorted(allowed_extensions))


    def check_index_status(self) -> None:
        if self.search_thread and self.search_thread.is_alive():
            self.index_status_var.set("Index status: wait until the current operation finishes, then click Check index.")
            return
        if self.index_status_thread and self.index_status_thread.is_alive():
            self.index_status_var.set("Index status: already checking. Please wait.")
            return

        roots = self._collect_selected_search_roots()
        allowed_extensions = self._collect_allowed_extensions()
        if not roots or not allowed_extensions:
            self.index_status_var.set("Index status: select at least one valid folder and file type.")
            return

        self.index_status_cancel_event.set()
        self.index_status_cancel_event = threading.Event()
        cancel_event = self.index_status_cancel_event
        self.index_status_var.set("Index status: checking for new or changed files...")

        def worker() -> None:
            try:
                summary = self.index_manager.get_stale_index_summary(
                    roots=roots,
                    allowed_extensions=allowed_extensions,
                    cancel_event=cancel_event,
                )
                if not cancel_event.is_set():
                    self.queue.put(("index_status", summary))
            except Exception as exc:
                if not cancel_event.is_set():
                    self.queue.put(("index_status_error", str(exc)))

        self.index_status_thread = threading.Thread(target=worker, daemon=True)
        self.index_status_thread.start()

    def _format_index_status_summary(self, summary: Dict[str, object]) -> str:
        files_checked = int(summary.get("files_checked", 0))
        indexed_files = int(summary.get("indexed_files", 0))
        new_files = int(summary.get("new_files", 0))
        changed_files = int(summary.get("changed_files", 0))
        deleted_files = int(summary.get("deleted_files", 0))
        missing_databases = int(summary.get("missing_databases", 0))
        invalid_databases = int(summary.get("invalid_databases", 0))
        errors = int(summary.get("errors", 0))
        stale_total = new_files + changed_files + deleted_files
        if missing_databases or invalid_databases:
            parts = []
            if missing_databases:
                parts.append(f"{missing_databases} missing {pluralize(missing_databases, 'index', 'indexes')}")
            if invalid_databases:
                parts.append(f"{invalid_databases} invalid {pluralize(invalid_databases, 'index', 'indexes')}")
            return (
                "Index status: update recommended — "
                + ", ".join(parts)
                + f"; {files_checked} searchable {pluralize(files_checked, 'file')} found."
                + " Click Archive settings at the top right, then click Update index to build or replace the affected index files."
            )
        if stale_total:
            details = []
            if new_files:
                details.append(f"{new_files} new")
            if changed_files:
                details.append(f"{changed_files} changed")
            if deleted_files:
                details.append(f"{deleted_files} deleted/missing")
            suffix = f"; {errors} scan {pluralize(errors, 'error')}" if errors else ""
            return (
                "Index status: update recommended — "
                + ", ".join(details)
                + f" since the last index update."
                + f" Click Archive settings at the top right, then click Update index to update the saved index{suffix}."
            )
        suffix = f"; {errors} scan {pluralize(errors, 'error')}" if errors else ""
        return (
            f"Index status: up to date for {indexed_files} indexed {pluralize(indexed_files, 'file')} "
            f"({files_checked} current {pluralize(files_checked, 'file')} checked{suffix})."
        )

    def _on_window_configure(self, event=None) -> None:
        self._update_status_wraplengths()

    def _update_status_wraplengths(self) -> None:
        wraplength = max(900, self.root.winfo_width() - 80)
        for widget in (
            getattr(self, "index_info_label", None),
            getattr(self, "index_status_label", None),
            getattr(self, "status_label", None),
        ):
            if widget is not None:
                widget.configure(wraplength=wraplength)

    def _browse_folder(self, variable: tk.StringVar, title: str) -> None:
        current = variable.get().strip()
        initial_dir = current if current and Path(current).exists() else str(self.app_folder)
        selected = filedialog.askdirectory(title=title, initialdir=initial_dir)
        if selected:
            variable.set(selected)

    def _update_folder_states(self) -> None:
        primary_enabled = self.use_primary_var.get()
        secondary_enabled = self.use_secondary_var.get()
        primary_state = "normal" if primary_enabled else "disabled"
        secondary_state = "normal" if secondary_enabled else "disabled"

        self.primary_folder_entry.config(state=primary_state)
        self.primary_browse_button.config(state=primary_state)
        self.primary_subfolders_check.config(state=primary_state)
        self.primary_check.config(style="FolderOn.TCheckbutton" if primary_enabled else "FolderOff.TCheckbutton")

        self.secondary_folder_entry.config(state=secondary_state)
        self.secondary_browse_button.config(state=secondary_state)
        self.secondary_subfolders_check.config(state=secondary_state)
        self.secondary_check.config(style="FolderOn.TCheckbutton" if secondary_enabled else "FolderOff.TCheckbutton")
        self._refresh_index_info()

    def _set_search_running_ui(self, running: bool) -> None:
        if running:
            self.search_button.config(state="disabled")
            self.update_button.config(state="disabled")
            self.check_index_button.config(state="disabled")
            self.cancel_button.config(state="normal")
            self.primary_check.config(state="disabled")
            self.secondary_check.config(state="disabled")
            self.primary_folder_entry.config(state="disabled")
            self.secondary_folder_entry.config(state="disabled")
            self.primary_browse_button.config(state="disabled")
            self.secondary_browse_button.config(state="disabled")
            self.primary_subfolders_check.config(state="disabled")
            self.secondary_subfolders_check.config(state="disabled")
            self.workbooks_check.config(state="disabled")
            self.text_documents_check.config(state="disabled")
            self.search_entry.config(state="disabled")
            self.progress_bar.grid()
            self.progress_bar.start(12)
        else:
            self.progress_bar.stop()
            self.progress_bar.grid_remove()
            self.search_button.config(state="normal")
            self.update_button.config(state="normal")
            self.check_index_button.config(state="normal")
            self.cancel_button.config(state="disabled")
            self.primary_check.config(state="normal")
            self.secondary_check.config(state="normal")
            self.workbooks_check.config(state="normal")
            self.text_documents_check.config(state="normal")
            self.search_entry.config(state="normal")
            if not self.search_var.get() and self.root.focus_get() != self.search_entry:
                self._show_search_placeholder()
            self._update_folder_states()

    def _set_details_editable(self, editable: bool) -> None:
        self.details_text.config(state="normal" if editable else "disabled")

    def _clear_details(self) -> None:
        self.detail_title_var.set("Result details")
        self.detail_meta_var.set("Select a result to review the evidence found in the archive.")
        self._set_details_editable(True)
        self.details_text.configure(wrap="word")
        try:
            self.details_text.delete("1.0", "end")
            self.details_text.insert("end", "Ready to search\n", ("empty_title",))
            self.details_text.insert(
                "end",
                "Search results will appear on the left. Select a file to see the matching rows, worksheet names, pages, and document text here.\n\n",
                ("empty_body",),
            )
            self.details_text.insert("end", "• Use All words when every term must be present.\n", ("empty_bullet",))
            self.details_text.insert("end", "• Use quotes, e.g. \"Sigma 5E\", for exact matches.\n", ("empty_bullet",))
            self.details_text.insert("end", "• Click Check index if results look out of date.\n", ("empty_bullet",))
        finally:
            self._set_details_editable(False)

    def _on_search_entry_focus_changed(self, event=None) -> None:
        """Show placeholder guidance only when the empty search box is not focused."""
        if event is not None and getattr(event, "type", None) == tk.EventType.FocusIn:
            self._hide_search_placeholder()
            return
        if not self.search_var.get().strip():
            self._show_search_placeholder()

    def _focus_empty_search_entry(self, event=None) -> str:
        """Move focus into the empty search entry without letting the placeholder behave like editable text."""
        try:
            if str(self.search_entry.cget("state")) == "disabled":
                return "break"
        except tk.TclError:
            pass
        self._hide_search_placeholder()
        self.search_entry.focus_set()
        self.search_entry.icursor(0)
        return "break"

    def _on_search_text_changed(self) -> None:
        """Keep the placeholder as a separate overlay, never as editable entry content."""
        if self.search_var.get():
            self._hide_search_placeholder()
        elif self.root.focus_get() != self.search_entry:
            self._show_search_placeholder()

    def _show_search_placeholder(self) -> None:
        """Display placeholder guidance over the empty entry without inserting it into the field value."""
        if self.search_placeholder_active or self.search_var.get():
            return
        self.search_placeholder_active = True
        self.search_placeholder_label.place(x=12, rely=0.5, anchor="w")
        self.search_placeholder_label.lift()

    def _hide_search_placeholder(self) -> None:
        """Hide the overlay placeholder while the user is typing or editing real text."""
        if not self.search_placeholder_active:
            return
        self.search_placeholder_active = False
        try:
            self.search_placeholder_label.place_forget()
        except tk.TclError:
            pass

    def _get_search_input(self) -> str:
        """Return only real user-entered search text."""
        return self.search_var.get().strip()

    def _highlight_details(self, start_index: str, end_index: str) -> None:
        if not self.current_highlight_terms:
            return
        for term in self.current_highlight_terms:
            if not term:
                continue
            search_start = start_index
            while True:
                match_index = self.details_text.search(term, search_start, stopindex=end_index, nocase=True)
                if not match_index:
                    break
                match_end = f"{match_index}+{len(term)}c"
                self.details_text.tag_add(HIGHLIGHT_TAG_NAME, match_index, match_end)
                search_start = match_end

    def _write_detail_text(self, text: str, tags: Tuple[str, ...] = ()) -> None:
        if not text:
            return
        start_index = self.details_text.index("end-1c")
        self.details_text.insert("end", text, tags)
        end_index = self.details_text.index("end-1c")
        if self.current_highlight_terms:
            self._highlight_details(start_index, end_index)

    def _write_result_folder_path(self, result: SearchResult) -> None:
        folder_path = str(Path(result.path).parent)
        self._write_detail_text(f"Folder: {folder_path}\n\n", ("muted",))

    @staticmethod
    def _split_workbook_row(row_text: str) -> List[str]:
        """Split an indexed worksheet row into display columns."""
        # Worksheet rows are indexed as cell text joined by " | ". Keep empty
        # trailing/leading text out of the display while preserving internal cells.
        return [part.strip() for part in row_text.split(" | ")]

    @staticmethod
    def _split_workbook_location(location_label: str) -> Tuple[str, str]:
        """Return sheet name and row number from a workbook location label."""
        parts = [part.strip() for part in location_label.split("|")]
        sheet_name = parts[0] if parts and parts[0] else "Sheet"
        row_number = ""
        for part in parts[1:]:
            if part.lower().startswith("row"):
                row_number = part[3:].strip()
                break
        return sheet_name, row_number

    @staticmethod
    def _excel_column_label(column_number: int) -> str:
        """Return Excel-style column labels: 1 -> A, 27 -> AA."""
        if column_number < 1:
            return "?"
        label = ""
        while column_number:
            column_number, remainder = divmod(column_number - 1, 26)
            label = chr(65 + remainder) + label
        return label

    @staticmethod
    def _trim_trailing_empty_cells(values: Sequence[str]) -> List[str]:
        """Keep leading/internal blank cells but remove unused trailing cells."""
        trimmed = list(values)
        while trimmed and not trimmed[-1]:
            trimmed.pop()
        return trimmed

    @staticmethod
    def _normalise_row_values(values: Sequence[object]) -> List[str]:
        """Convert workbook row values into display strings while preserving column positions."""
        return ArchiveSearchApp._trim_trailing_empty_cells([display_text(value) for value in values])

    @staticmethod
    def _valid_row_number(row_number_text: str) -> Optional[int]:
        try:
            row_number = int(str(row_number_text).strip())
        except (TypeError, ValueError):
            return None
        return row_number if row_number > 0 else None

    def _choose_header_from_candidate_rows(self, candidate_rows: Sequence[Tuple[int, Sequence[object]]]) -> Tuple[List[str], Optional[int]]:
        """Pick the most likely header row above the matches.

        Sales archive workbooks usually have a single header row near the top of
        each sheet, but some contain a title row before the table. To keep the
        display useful, prefer the earliest row among the rows with the most
        populated cells. Blank/short title rows therefore lose to a fuller header
        row, while data rows later in the sheet are avoided.
        """
        best_values: List[str] = []
        best_row_number: Optional[int] = None
        best_non_empty_count = 0
        best_original_length = 0

        for row_number, raw_values in candidate_rows:
            values = self._normalise_row_values(raw_values)
            non_empty_count = sum(1 for value in values if value)
            if not non_empty_count:
                continue
            original_length = len(values)
            if (
                non_empty_count > best_non_empty_count
                or (non_empty_count == best_non_empty_count and best_row_number is not None and row_number < best_row_number)
                or (non_empty_count == best_non_empty_count and best_row_number is None)
            ):
                best_values = values
                best_row_number = row_number
                best_non_empty_count = non_empty_count
                best_original_length = original_length
            elif non_empty_count == best_non_empty_count and original_length > best_original_length and best_row_number == row_number:
                best_values = values
                best_original_length = original_length

        return best_values, best_row_number

    def _read_openpyxl_sheet_display_data(self, workbook, sheet_name: str, row_numbers: Sequence[int]) -> Tuple[List[str], Optional[int], Dict[int, List[str]]]:
        ws = workbook[sheet_name]
        valid_rows = sorted({row_number for row_number in row_numbers if row_number > 0})
        if not valid_rows:
            return [], None, {}

        first_match_row = min(valid_rows)
        header_candidates: List[Tuple[int, Sequence[object]]] = []
        if first_match_row > 1:
            max_header_scan_row = min(first_match_row - 1, 25)
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_header_scan_row, values_only=True), start=1):
                header_candidates.append((idx, row))
        headers, header_row_number = self._choose_header_from_candidate_rows(header_candidates)

        rows: Dict[int, List[str]] = {}
        for row_number in valid_rows:
            try:
                row_values = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
            except StopIteration:
                continue
            rows[row_number] = self._normalise_row_values(row_values)
        return headers, header_row_number, rows

    def _read_xlrd_sheet_display_data(self, book, sheet_name: str, row_numbers: Sequence[int]) -> Tuple[List[str], Optional[int], Dict[int, List[str]]]:
        sheet = book.sheet_by_name(sheet_name)
        valid_rows = sorted({row_number for row_number in row_numbers if row_number > 0 and row_number <= sheet.nrows})
        if not valid_rows:
            return [], None, {}

        def get_row_values(one_based_row_number: int) -> List[object]:
            zero_based = one_based_row_number - 1
            raw_values = sheet.row_values(zero_based)
            raw_types = sheet.row_types(zero_based)
            return [convert_xls_value(value, ctype, book.datemode) for value, ctype in zip(raw_values, raw_types)]

        first_match_row = min(valid_rows)
        header_candidates: List[Tuple[int, Sequence[object]]] = []
        if first_match_row > 1:
            max_header_scan_row = min(first_match_row - 1, 25)
            for row_number in range(1, max_header_scan_row + 1):
                header_candidates.append((row_number, get_row_values(row_number)))
        headers, header_row_number = self._choose_header_from_candidate_rows(header_candidates)
        rows = {row_number: self._normalise_row_values(get_row_values(row_number)) for row_number in valid_rows}
        return headers, header_row_number, rows

    def _read_xlsb_sheet_display_data(self, workbook, sheet_name: str, row_numbers: Sequence[int]) -> Tuple[List[str], Optional[int], Dict[int, List[str]]]:
        valid_rows = sorted({row_number for row_number in row_numbers if row_number > 0})
        if not valid_rows:
            return [], None, {}

        first_match_row = min(valid_rows)
        max_header_scan_row = min(first_match_row - 1, 25) if first_match_row > 1 else 0
        wanted_rows = set(valid_rows)
        header_candidates: List[Tuple[int, Sequence[object]]] = []
        rows: Dict[int, List[str]] = {}

        with workbook.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                if not row:
                    continue
                row_number = int(getattr(row[0], "r", 0) or 0)
                if not row_number:
                    continue
                raw_values = [getattr(cell, "v", None) for cell in row]
                if 1 <= row_number <= max_header_scan_row:
                    header_candidates.append((row_number, raw_values))
                if row_number in wanted_rows:
                    rows[row_number] = self._normalise_row_values(raw_values)
                if row_number > max(max_header_scan_row, max(wanted_rows, default=0)):
                    break

        headers, header_row_number = self._choose_header_from_candidate_rows(header_candidates)
        return headers, header_row_number, rows

    def _read_workbook_display_data(self, result: SearchResult, grouped_rows: Dict[str, List[Tuple[str, List[str]]]]) -> Tuple[Dict[str, Dict[str, object]], List[str]]:
        """Read live workbook headers/rows and sheet order for the details pane."""
        file_path = Path(result.path)
        if not file_path.exists():
            return {}, []
        sheet_row_numbers: Dict[str, List[int]] = {}
        for sheet_name, rows in grouped_rows.items():
            numbers = [number for number in (self._valid_row_number(row_number) for row_number, _cells in rows) if number]
            if numbers:
                sheet_row_numbers[sheet_name] = numbers
        if not sheet_row_numbers:
            return {}, []

        suffix = file_path.suffix.lower()
        data: Dict[str, Dict[str, object]] = {}
        workbook_sheet_order: List[str] = []

        try:
            if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
                try:
                    workbook_sheet_order = list(workbook.sheetnames)
                    for sheet_name, row_numbers in sheet_row_numbers.items():
                        if sheet_name not in workbook.sheetnames:
                            continue
                        headers, header_row_number, rows = self._read_openpyxl_sheet_display_data(workbook, sheet_name, row_numbers)
                        data[sheet_name] = {"headers": headers, "header_row_number": header_row_number, "rows": rows}
                finally:
                    workbook.close()
            elif suffix == ".xls" and HAS_XLRD:
                book = xlrd.open_workbook(str(file_path), on_demand=True)
                try:
                    workbook_sheet_order = list(book.sheet_names())
                    available_sheets = set(workbook_sheet_order)
                    for sheet_name, row_numbers in sheet_row_numbers.items():
                        if sheet_name not in available_sheets:
                            continue
                        headers, header_row_number, rows = self._read_xlrd_sheet_display_data(book, sheet_name, row_numbers)
                        data[sheet_name] = {"headers": headers, "header_row_number": header_row_number, "rows": rows}
                finally:
                    try:
                        book.release_resources()
                    except Exception:
                        pass
            elif suffix == ".xlsb" and HAS_PYXLSB and open_xlsb_workbook is not None:
                with open_xlsb_workbook(str(file_path)) as workbook:
                    workbook_sheet_order = list(workbook.sheets)
                    available_sheets = set(workbook_sheet_order)
                    for sheet_name, row_numbers in sheet_row_numbers.items():
                        if sheet_name not in available_sheets:
                            continue
                        headers, header_row_number, rows = self._read_xlsb_sheet_display_data(workbook, sheet_name, row_numbers)
                        data[sheet_name] = {"headers": headers, "header_row_number": header_row_number, "rows": rows}
        except Exception:
            # Keep the UI responsive and still useful if a workbook is locked,
            # moved, protected, or uses an optional format reader that is missing.
            return {}, []

        return data, workbook_sheet_order

    def _build_workbook_headings(self, headers: Sequence[str], max_excel_columns: int) -> List[str]:
        """Use workbook header names where available and Excel letters as safe fallbacks."""
        headings = ["Row"]
        seen: Dict[str, int] = {}
        for column_index in range(1, max_excel_columns + 1):
            raw_heading = headers[column_index - 1] if column_index - 1 < len(headers) else ""
            heading = str(raw_heading).strip() or self._excel_column_label(column_index)
            lower_heading = heading.casefold()
            if lower_heading in seen:
                seen[lower_heading] += 1
                heading = f"{heading} ({self._excel_column_label(column_index)})"
            else:
                seen[lower_heading] = 1
            headings.append(heading)
        return headings

    def _build_workbook_sheet_sections(self, result: SearchResult) -> List[Tuple[str, int, str, str, List[str]]]:
        """Group workbook matches by worksheet and format each section as an aligned table."""
        grouped_rows: Dict[str, List[Tuple[str, List[str]]]] = {}
        sheet_order: List[str] = []

        for snippet in result.snippets:
            sheet_name, row_number = self._split_workbook_location(snippet.location_label)
            cells = self._split_workbook_row(snippet.preview)
            if sheet_name not in grouped_rows:
                grouped_rows[sheet_name] = []
                sheet_order.append(sheet_name)
            grouped_rows[sheet_name].append((row_number or "?", cells))

        live_workbook_data, workbook_sheet_order = self._read_workbook_display_data(result, grouped_rows)
        workbook_ordered_sheets = [sheet_name for sheet_name in workbook_sheet_order if sheet_name in grouped_rows]
        fallback_sheets = [sheet_name for sheet_name in sheet_order if sheet_name not in workbook_ordered_sheets]
        display_sheet_order = workbook_ordered_sheets + fallback_sheets

        sections: List[Tuple[str, int, str, str, List[str]]] = []
        for sheet_name in display_sheet_order:
            rows = sorted(
                grouped_rows[sheet_name],
                key=lambda item: (
                    self._valid_row_number(item[0]) is None,
                    self._valid_row_number(item[0]) or 0,
                    str(item[0]),
                ),
            )
            sheet_data = live_workbook_data.get(sheet_name, {})
            live_rows = sheet_data.get("rows", {}) if isinstance(sheet_data, dict) else {}
            headers = sheet_data.get("headers", []) if isinstance(sheet_data, dict) else []
            header_row_number = sheet_data.get("header_row_number") if isinstance(sheet_data, dict) else None
            if not isinstance(headers, list):
                headers = []
            if not isinstance(live_rows, dict):
                live_rows = {}
            if not isinstance(header_row_number, int):
                header_row_number = None

            table_rows: List[List[str]] = []
            max_excel_columns = len(headers)
            for row_number_text, fallback_cells in rows:
                row_number = self._valid_row_number(row_number_text)
                row_cells = live_rows.get(row_number, fallback_cells) if row_number is not None else fallback_cells
                row_cells = self._trim_trailing_empty_cells([str(cell) for cell in row_cells])
                max_excel_columns = max(max_excel_columns, len(row_cells))
                table_rows.append([row_number_text or "?", *row_cells])

            headings = self._build_workbook_headings(headers, max_excel_columns)
            max_columns = len(headings)

            widths: List[int] = []
            for column_index in range(max_columns):
                values = [headings[column_index]]
                values.extend(row[column_index] if column_index < len(row) else "" for row in table_rows)
                widths.append(max(len(value) for value in values))

            def format_line(values: Sequence[str]) -> str:
                padded = [
                    (values[column_index] if column_index < len(values) else "").ljust(widths[column_index])
                    for column_index in range(max_columns)
                ]
                return " | ".join(padded)

            header_line = format_line(headings).rstrip()
            rule_line = "-+-".join("-" * width for width in widths).rstrip()
            row_lines = [format_line(row) for row in table_rows]
            sections.append((sheet_name, len(rows), header_line, rule_line, row_lines))

        return sections

    def _show_result_details(self, result: SearchResult) -> None:
        type_label = infer_type_label(result.file_type)
        self.detail_title_var.set(result.document_name)
        match_label = f"{result.match_count} {pluralize(result.match_count, 'match')}"
        meta = f"{type_label} | {match_label} | Archive: {result.root_label}"
        if result.file_type == "excel":
            meta += f" | {result.match_count} {pluralize(result.match_count, 'worksheet row match', 'worksheet row matches')}"
        elif result.first_location_label:
            meta += f" | First match: {result.first_location_label}"
        self.detail_meta_var.set(meta)
        self._set_details_editable(True)
        try:
            self.details_text.delete("1.0", "end")
            self._write_result_folder_path(result)
            if result.file_type == "excel":
                self.details_text.configure(wrap="none")
                shown = len(result.snippets)
                self._write_detail_text(
                    f"Workbook result: {result.match_count} {pluralize(result.match_count, 'matching row', 'matching rows')} found.\n",
                    ("heading",),
                )
                if shown and shown < result.match_count:
                    self._write_detail_text(
                        f"Showing {shown} of {result.match_count} rows. ",
                        ("muted",),
                    )
                self._write_detail_text(
                    "Matches are grouped by sheet. Column headings are taken from the workbook where available; blank headings use Excel letters.\n\n",
                    ("muted",),
                )
                for sheet_name, sheet_match_count, header_line, rule_line, row_lines in self._build_workbook_sheet_sections(result):
                    self._write_detail_text(
                        f"{sheet_name} — {sheet_match_count} {pluralize(sheet_match_count, 'matching row', 'matching rows')}\n",
                        ("workbook_sheet",),
                    )
                    self._write_detail_text(header_line + "\n", ("workbook_header",))
                    self._write_detail_text(rule_line + "\n", ("workbook_rule",))
                    for row_index, row_line in enumerate(row_lines):
                        row_tag = "workbook_row_even" if row_index % 2 else "workbook_row_odd"
                        self._write_detail_text(row_line + "\n", (row_tag,))
                    self._write_detail_text("\n")
            else:
                self.details_text.configure(wrap="word")
                full_text = self.index_manager.get_indexed_document_text(result)
                if full_text:
                    self._write_detail_text(full_text.rstrip() + "\n")
                else:
                    for snippet in result.snippets:
                        self._write_detail_text(f"{snippet.location_label}\n", ("snippet_label",))
                        self._write_detail_text(f"{snippet.preview}\n\n")
            self.details_text.see("1.0")
        finally:
            self._set_details_editable(False)

    def _column_from_tree_identifier(self, column_identifier: str) -> Optional[str]:
        """Translate a Treeview column identifier such as '#1' to our column key."""
        if not column_identifier or not column_identifier.startswith("#"):
            return None
        try:
            column_index = int(column_identifier[1:]) - 1
        except ValueError:
            return None
        columns = tuple(self.results_tree["columns"])
        if 0 <= column_index < len(columns):
            return str(columns[column_index])
        return None

    def _auto_size_result_column(self, column: str) -> None:
        """Fit a result-table column to the visible heading and row values."""
        if column not in self.result_heading_labels:
            return
        try:
            value_font = tkfont.Font(font=self._ui_font("body", 9))
            heading_font = tkfont.Font(font=self._ui_font("body", 9, "bold"))
        except tk.TclError:
            return
        heading = self.result_heading_labels.get(column, column)
        width = heading_font.measure(heading) + 32
        for item_id in self.results_tree.get_children(""):
            text = str(self.results_tree.set(item_id, column))
            width = max(width, value_font.measure(text) + 32)
        min_width = int(self.results_tree.column(column, option="minwidth") or 60)
        # Keep the table usable even when a file name is exceptionally long.
        max_width = 900 if column == "document" else 220
        self.results_tree.column(column, width=max(min_width, min(width, max_width)))

    def _on_results_double_click(self, event=None):
        """Auto-fit headings/dividers, or open a result when its row is double-clicked."""
        if event is None:
            return "break"
        region = self.results_tree.identify_region(event.x, event.y)
        if region in {"heading", "separator"}:
            column = self._column_from_tree_identifier(self.results_tree.identify_column(event.x))
            if column:
                self._auto_size_result_column(column)
            return "break"
        if region in {"cell", "tree"}:
            item_id = self.results_tree.identify_row(event.y)
            if item_id:
                self.results_tree.selection_set(item_id)
                self.results_tree.focus(item_id)
                self._on_result_selected()
                self.open_selected_result()
            return "break"
        return "break"

    def _build_results_context_menu(self) -> None:
        """Build a right-click menu for the most common result actions."""
        self.results_context_menu = tk.Menu(
            self.root,
            tearoff=0,
            bg=BRAND_WHITE,
            fg=BRAND_BLACK,
            activebackground=SOFT_BLUE,
            activeforeground=BRAND_BLACK,
            font=self._ui_font("body", 9),
        )
        self.results_context_menu.add_command(label="Open selected", command=self.open_selected_result)
        self.results_context_menu.add_command(label="Show in folder", command=self.reveal_selected_result)
        self.results_context_menu.add_separator()
        self.results_context_menu.add_command(label="Copy file path", command=self.copy_selected_path)

    def _show_results_context_menu(self, event=None):
        """Select the row under the pointer and show contextual actions."""
        if event is None:
            return "break"
        item_id = self.results_tree.identify_row(event.y)
        if not item_id:
            return "break"
        self.results_tree.selection_set(item_id)
        self.results_tree.focus(item_id)
        self._on_result_selected()
        try:
            self.results_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.results_context_menu.grab_release()
        return "break"

    def _update_result_headings(self) -> None:
        """Render clickable result-table headings and show the active sort direction."""
        for column, label in self.result_heading_labels.items():
            suffix = ""
            if self.result_sort_column == column:
                suffix = " ↓" if self.result_sort_reverse else " ↑"
            self.results_tree.heading(
                column,
                text=label + suffix,
                command=lambda selected_column=column: self._sort_results_by_column(selected_column),
            )

    def _result_sort_key(self, item_id: str, column: str):
        value = self.results_tree.set(item_id, column)
        if column == "matches":
            try:
                return int(value)
            except (TypeError, ValueError):
                return 0
        return str(value).casefold()

    def _sort_results_by_column(self, column: str) -> None:
        """Toggle ascending/descending sort for the selected result-table column."""
        if column not in self.result_heading_labels:
            return
        if self.result_sort_column == column:
            self.result_sort_reverse = not self.result_sort_reverse
        else:
            self.result_sort_column = column
            self.result_sort_reverse = False

        selected = self.results_tree.selection()
        focused = self.results_tree.focus()
        items = list(self.results_tree.get_children(""))
        items.sort(key=lambda item_id: self._result_sort_key(item_id, column), reverse=self.result_sort_reverse)
        for index, item_id in enumerate(items):
            self.results_tree.move(item_id, "", index)
        self._update_result_headings()
        self._retag_result_rows()
        if selected:
            self.results_tree.selection_set(selected)
        if focused:
            self.results_tree.focus(focused)

    def _retag_result_rows(self) -> None:
        """Apply alternating row backgrounds and a clear hover highlight."""
        for index, item_id in enumerate(self.results_tree.get_children("")):
            if item_id == self.hovered_result_iid:
                # Use the hover tag on its own so it reliably overrides row banding
                # across the Tk themes used on different Windows installs.
                self.results_tree.item(item_id, tags=("hover",))
            else:
                self.results_tree.item(item_id, tags=("even" if index % 2 else "odd",))

    def _on_results_mouse_motion(self, event=None) -> None:
        """Highlight the result row under the pointer."""
        if event is None:
            return
        region = self.results_tree.identify_region(event.x, event.y)
        item_id = ""
        if region not in {"heading", "separator", "nothing"}:
            item_id = self.results_tree.identify_row(event.y)
        item_id = item_id or None
        if item_id != self.hovered_result_iid:
            self.hovered_result_iid = item_id
            self._retag_result_rows()

    def _on_results_mouse_leave(self, event=None) -> None:
        if self.hovered_result_iid is not None:
            self.hovered_result_iid = None
            self._retag_result_rows()


    def _update_results_summary(self) -> None:
        result_count = len(self.results_by_iid)
        total_matches = sum(result.match_count for result in self.results_by_iid.values())
        if not result_count:
            self.result_summary_var.set("No results yet")
            return
        self.result_summary_var.set(
            f"{result_count} matching {pluralize(result_count, 'file')} • {total_matches} total {pluralize(total_matches, 'match')}"
        )

    def _update_selection_actions(self) -> None:
        state = "normal" if self.results_tree.selection() else "disabled"
        for button_name in ("open_button", "show_folder_button", "copy_path_button"):
            button = getattr(self, button_name, None)
            if button is not None:
                button.config(state=state)

    def _get_selected_result(self) -> Optional[SearchResult]:
        selected = self.results_tree.selection()
        if not selected:
            return None
        return self.results_by_iid.get(selected[0])

    def _clear_results_tree(self) -> None:
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        self.results_by_iid.clear()
        self.hovered_result_iid = None
        self._update_results_summary()
        self._update_selection_actions()

    def _append_result_to_tree(self, result: SearchResult) -> None:
        iid = self.results_tree.insert(
            "",
            "end",
            values=(
                result.document_name,
                infer_type_label(result.file_type),
                result.match_count,
            ),
        )
        self.results_by_iid[iid] = result
        if self.result_sort_column:
            items = list(self.results_tree.get_children(""))
            items.sort(
                key=lambda item_id: self._result_sort_key(item_id, self.result_sort_column or "document"),
                reverse=self.result_sort_reverse,
            )
            for index, item_id in enumerate(items):
                self.results_tree.move(item_id, "", index)
        self._retag_result_rows()
        self._update_results_summary()

    def _on_result_selected(self, event=None) -> None:
        selected = self.results_tree.selection()
        if not selected:
            self._update_selection_actions()
            return
        result = self.results_by_iid.get(selected[0])
        if result is not None:
            self._show_result_details(result)
        self._update_selection_actions()

    def open_selected_result(self) -> None:
        result = self._get_selected_result()
        if result is None:
            messagebox.showinfo("No selection", "Please select a result first.")
            return
        FileOpener.open_result(
            file_type=result.file_type,
            file_path=result.path,
            sheet_name=result.open_sheet,
            row_number=result.open_row_number,
        )

    def reveal_selected_result(self) -> None:
        result = self._get_selected_result()
        if result is None:
            messagebox.showinfo("No selection", "Please select a result first.")
            return
        FileOpener.reveal_in_file_manager(result.path)

    def copy_selected_path(self) -> None:
        result = self._get_selected_result()
        if result is None:
            messagebox.showinfo("No selection", "Please select a result first.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(result.path)
        self.status_var.set("Copied selected result path to clipboard.")

    def clear_output(self) -> None:
        self._clear_results_tree()
        self._clear_details()
        self.status_var.set("Ready.")

    def _refresh_index_info(self) -> None:
        roots = []
        if self.use_primary_var.get():
            primary_text = self.primary_folder_var.get().strip()
            if primary_text:
                roots.append(SearchRoot("Primary", Path(primary_text).expanduser(), self.primary_subfolders_var.get()))
        if self.use_secondary_var.get():
            secondary_text = self.secondary_folder_var.get().strip()
            if secondary_text:
                roots.append(SearchRoot("Secondary", Path(secondary_text).expanduser(), self.secondary_subfolders_var.get()))
        allowed_extensions = self._collect_allowed_extensions()
        self.index_info_var.set(self.index_manager.describe_selected_index_paths(tuple(roots), allowed_extensions))

    def _get_selected_search_roots(self) -> Optional[Tuple[SearchRoot, ...]]:
        roots: List[SearchRoot] = []
        if not self.use_primary_var.get() and not self.use_secondary_var.get():
            messagebox.showwarning("No Search Folder Selected", "Please enable the primary folder, the secondary folder, or both.")
            return None
        if self.use_primary_var.get():
            primary_text = self.primary_folder_var.get().strip()
            if not primary_text:
                messagebox.showwarning("Missing Folder", "The primary folder is enabled, but no primary folder was entered.")
                return None
            primary_path = Path(primary_text).expanduser()
            if not primary_path.exists() or not primary_path.is_dir():
                messagebox.showerror("Invalid Primary Folder", f"The primary search folder does not exist or is not a folder:\n{primary_path}")
                return None
            roots.append(SearchRoot("Primary", primary_path, self.primary_subfolders_var.get()))
        if self.use_secondary_var.get():
            secondary_text = self.secondary_folder_var.get().strip()
            if not secondary_text:
                messagebox.showwarning("Missing Folder", "The secondary folder is enabled, but no secondary folder was entered.")
                return None
            secondary_path = Path(secondary_text).expanduser()
            if not secondary_path.exists() or not secondary_path.is_dir():
                messagebox.showerror("Invalid Secondary Folder", f"The secondary search folder does not exist or is not a folder:\n{secondary_path}")
                return None
            roots.append(SearchRoot("Secondary", secondary_path, self.secondary_subfolders_var.get()))
        return tuple(roots)

    def _get_allowed_extensions(self) -> Optional[Tuple[str, ...]]:
        allowed_extensions = self._collect_allowed_extensions()
        if not allowed_extensions:
            messagebox.showwarning("No File Types Selected", "Please tick Workbooks, Text documents, or both.")
            return None
        return allowed_extensions

    @staticmethod
    def _format_blocking_file_type_issues(issues: Sequence[Dict[str, str]]) -> str:
        shown = []
        for index, issue in enumerate(issues[:10], start=1):
            shown.append(
                f"Folder:\n{issue.get('folder', '')}\n\n"
                f"File:\n{issue.get('file_name', '')}\n\n"
                f"Extension suggests: {issue.get('expected', '')}\n"
                f"Actual file content: {issue.get('actual', '')}"
            )
        extra = ""
        if len(issues) > 10:
            extra = f"\n\nOnly the first 10 issues are shown. There are {len(issues)} issues in total."
        return (
            "Update index has not started because one or more files have an extension that does not match the file content.\n\n"
            + "\n\n".join(shown)
            + extra
            + "\n\nRename, remove, or replace the problem file(s), then click Archive settings at the top right, then click Update index again."
        )

    def _stop_index_status_check(self) -> None:
        """Stop any in-progress manual index-status check before changing an index file."""
        self.index_status_cancel_event.set()
        if self.index_status_thread and self.index_status_thread.is_alive():
            self.index_status_thread.join(timeout=2.0)


    def _start_background_job(
        self,
        *,
        search_terms: Sequence[str],
        update_only: bool = False,
    ) -> None:
        """Start a search or incremental update in a daemon worker thread."""
        if self.search_thread and self.search_thread.is_alive():
            return

        search_roots = self._get_selected_search_roots()
        if not search_roots:
            return
        allowed_extensions = self._get_allowed_extensions()
        if not allowed_extensions:
            return

        if update_only:
            self._stop_index_status_check()
            self.status_var.set("Checking file types before updating the index...")
            self.root.update_idletasks()
            blocking_issues = self.index_manager.find_blocking_file_type_issues(search_roots, allowed_extensions)
            if blocking_issues:
                self.status_var.set("Update cancelled: fix the file type issue(s), then try again.")
                messagebox.showerror(
                    "File Type Issue Found",
                    self._format_blocking_file_type_issues(blocking_issues),
                )
                return

        if update_only:
            build_requirements = self.index_manager.get_index_build_requirements(search_roots)
            if build_requirements:
                missing_lines = []
                invalid_lines = []
                for item in build_requirements:
                    root = item["root"]
                    database_path = item["database_path"]
                    line = f"{root.label}: {root.path}\n\n{database_path.name}\n"
                    if item.get("reason") == "invalid":
                        invalid_lines.append(line)
                    else:
                        missing_lines.append(line)

                message_parts = []
                if missing_lines:
                    message_parts.append(
                        "The following selected folder(s) do not already have a saved Archive Search index:\n\n"
                        + "\n".join(missing_lines)
                    )
                if invalid_lines:
                    message_parts.append(
                        "The following selected folder(s) already have a file at the expected index path, "
                        "but it is not a valid Archive Search index. It will be overwritten with a fresh build:\n\n"
                        + "\n".join(invalid_lines)
                    )

                confirmed = messagebox.askyesno(
                    "Index Will Be Built First",
                    "\n\n".join(message_parts)
                    + "\nUpdate index will build the listed index file(s) from scratch, so it may take a while. "
                    "It will not work if another user or computer is currently editing or updating the same index. Continue?",
                )
                if not confirmed:
                    self.status_var.set("Update cancelled.")
                    return

        self.cancel_event.clear()
        if not update_only:
            self.index_status_cancel_event.set()
        self.current_job_mode = "update" if update_only else "search"
        self.current_highlight_terms = build_highlight_terms(search_terms)
        self.clear_output()
        self.status_var.set("Working...")
        self._set_search_running_ui(True)

        runner = SearchRunner(
            index_manager=self.index_manager,
            queue_out=self.queue,
            cancel_event=self.cancel_event,
        )
        self.search_thread = threading.Thread(
            target=runner.run,
            args=(search_roots, allowed_extensions, tuple(search_terms), self.match_mode_var.get()),
            kwargs={
                "update_only": update_only,
            },
            daemon=True,
        )
        self.search_thread.start()

    def start_search(self) -> None:
        if self.search_thread and self.search_thread.is_alive():
            return
        search_input = self._get_search_input()
        if not search_input:
            messagebox.showwarning("Missing Search Terms", "Please type one or more search terms.")
            return
        try:
            search_terms = tuple(term.lower() for term in shlex.split(search_input))
        except ValueError as exc:
            messagebox.showerror("Invalid Search Input", f"Could not parse search terms:\n{exc}")
            return
        self._start_background_job(search_terms=search_terms)

    def update_index(self) -> None:
        self._start_background_job(search_terms=(), update_only=True)

    def cancel_process(self) -> None:
        if self.search_thread and self.search_thread.is_alive():
            self.cancel_event.set()
            self.status_var.set("Cancelling...")
            self.cancel_button.config(state="disabled")

    def _poll_queue(self) -> None:
        pending_item_batches: List[List[Dict[str, object]]] = []
        latest_status: Optional[str] = None
        finished_payload: Optional[Dict[str, object]] = None
        fatal_payload: Optional[str] = None
        index_status_payload: Optional[Dict[str, object]] = None
        index_status_error: Optional[str] = None

        try:
            while True:
                item_type, payload = self.queue.get_nowait()
                if item_type == "items":
                    pending_item_batches.append(payload)
                elif item_type == "status":
                    latest_status = payload
                elif item_type == "finished":
                    finished_payload = payload
                elif item_type == "fatal":
                    fatal_payload = payload
                elif item_type == "index_status":
                    index_status_payload = payload
                elif item_type == "index_status_error":
                    index_status_error = payload
        except queue.Empty:
            pass

        if index_status_payload is not None:
            self._refresh_index_info()
            self.index_status_var.set(self._format_index_status_summary(index_status_payload))

        if index_status_error is not None:
            self.index_status_var.set(f"Index status: could not check index ({index_status_error}).")

        if pending_item_batches:
            for batch in pending_item_batches:
                for item in batch:
                    kind = item["kind"]
                    payload = item["payload"]
                    if kind == "result":
                        self._append_result_to_tree(payload)
            if not self.results_tree.selection() and self.results_tree.get_children():
                first = self.results_tree.get_children()[0]
                self.results_tree.selection_set(first)
                self.results_tree.focus(first)
                self.results_tree.see(first)
                self._on_result_selected()

        if latest_status is not None:
            self.status_var.set(latest_status)

        if finished_payload is not None:
            files_scanned = int(finished_payload.get("files_scanned", 0))
            indexed_files = int(finished_payload.get("indexed_files", 0))
            matched_files = int(finished_payload.get("matched_files", 0))
            total_matches = int(finished_payload.get("total_matches", 0))
            cancelled = bool(finished_payload.get("cancelled", False))
            elapsed_seconds = float(finished_payload.get("elapsed_seconds", 0.0))
            mode = str(finished_payload.get("mode", self.current_job_mode))
            files_reused = int(finished_payload.get("files_reused", 0))
            files_refreshed = int(finished_payload.get("files_refreshed", 0))
            files_deleted = int(finished_payload.get("files_deleted", 0))
            files_failed = int(finished_payload.get("files_failed", 0))
            first_error = str(finished_payload.get("first_error") or "")

            if cancelled:
                final_message = (
                    f"Cancelled. Scanned {files_scanned} {pluralize(files_scanned, 'file')}. "
                    f"Indexed set currently contains {indexed_files} {pluralize(indexed_files, 'file')}. "
                    f"Time: {elapsed_seconds:.2f} seconds."
                )
            elif mode == "update":
                warning = ""
                if files_failed:
                    warning = (
                        f" Warning: {files_failed} {pluralize(files_failed, 'file')} could not be indexed."
                    )
                    if first_error:
                        warning += f" First issue: {first_error}."
                final_message = (
                    f"Update finished. Checked {files_scanned} {pluralize(files_scanned, 'file')}. "
                    f"Kept {files_reused} unchanged, updated {files_refreshed}, removed {files_deleted}. "
                    f"The saved index now contains {indexed_files} {pluralize(indexed_files, 'file')}."
                    f"{warning} "
                    f"Time: {elapsed_seconds:.2f} seconds."
                )
            elif mode == "initial_index_required":
                missing_roots = str(finished_payload.get("missing_roots") or "the selected folder")
                final_message = (
                    f"Search not started. The saved index is missing or is not a valid Archive Search index for {missing_roots}. "
                    "Click Archive settings at the top right, then click Update index first. After it finishes, click Search archive again."
                )
            else:
                final_message = (
                    f"Finished searching the saved index. "
                    f"Indexed {indexed_files} {pluralize(indexed_files, 'file')} total. "
                    f"Found {matched_files} matching {pluralize(matched_files, 'file')} with {total_matches} total {pluralize(total_matches, 'match')}. "
                    f"Time: {elapsed_seconds:.2f} seconds."
                )

            self.status_var.set(final_message)
            self._set_search_running_ui(False)
            if not cancelled and mode == "update":
                self._refresh_index_info()
                self.index_status_var.set("Index status: up to date.")

        if fatal_payload is not None:
            self.status_var.set("Operation failed.")
            self._set_search_running_ui(False)
            messagebox.showerror("Operation failed", fatal_payload)

        self.root.after(100, self._poll_queue)
